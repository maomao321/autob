from __future__ import annotations

import csv
import hashlib
import json
import uuid
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from autoquant.config import default_config_path


OPENAI_FILES_URL = "https://api.openai.com/v1/files"
OPENAI_VECTOR_STORES_URL = "https://api.openai.com/v1/vector_stores"
MAX_HTTP_RESPONSE_BYTES = 2_000_000
MAX_IMPORT_FILE_BYTES = 50_000_000
MAX_IMPORT_ROWS = 500_000


class ExperienceError(RuntimeError):
    """A safe, user-displayable experience import or upload failure."""


@dataclass(frozen=True, slots=True)
class MarketBar:
    symbol: str
    timestamp_ms: int
    open: Decimal
    high: Decimal
    low: Decimal
    close: Decimal
    volume: Decimal = Decimal("0")
    interval: str = ""
    group_id: str = ""
    pattern_name: str = ""


@dataclass(frozen=True, slots=True)
class TradeExperience:
    experience_id: str
    record_type: str
    external_id: str
    symbol: str
    side: str
    outcome: str
    entry_time_ms: int
    exit_time_ms: int
    entry_time: str
    exit_time: str
    quantity: str
    entry_price: str
    exit_price: str
    entry_fee: str
    exit_fee: str
    gross_pnl: str
    net_pnl: str
    return_percent: str
    holding_seconds: int
    source: str
    pre_entry_pattern: dict[str, Any] = field(default_factory=dict)
    market_context: dict[str, str] = field(default_factory=dict)
    notes: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class ExperienceSummary:
    total: int
    trades: int
    patterns: int
    wins: int
    losses: int
    breakeven: int
    with_kline: int
    net_pnl: Decimal


@dataclass(frozen=True, slots=True)
class ExperienceImportResult:
    experiences: list[TradeExperience]
    trade_rows: int
    kline_rows: int


@dataclass(frozen=True, slots=True)
class UploadResult:
    vector_store_id: str
    file_id: str
    attachment_id: str
    status: str


_TRADE_ALIASES = {
    "trade_id": ("trade_id", "id", "交易编号", "交易id", "记录编号"),
    "symbol": ("symbol", "ticker", "code", "股票", "股票代码", "标的"),
    "side": ("side", "direction", "方向", "交易方向", "多空"),
    "entry_time": (
        "entry_time",
        "open_time",
        "buy_time",
        "入场时间",
        "开仓时间",
        "买入时间",
    ),
    "exit_time": (
        "exit_time",
        "close_time",
        "sell_time",
        "出场时间",
        "平仓时间",
        "卖出时间",
    ),
    "entry_price": (
        "entry_price",
        "open_price",
        "buy_price",
        "入场价",
        "开仓价",
        "买入价",
    ),
    "exit_price": (
        "exit_price",
        "close_price",
        "sell_price",
        "出场价",
        "平仓价",
        "卖出价",
    ),
    "quantity": ("quantity", "qty", "size", "数量", "成交数量", "仓位"),
    "fee": ("fee", "total_fee", "commission", "手续费", "总手续费"),
    "entry_fee": ("entry_fee", "open_fee", "开仓手续费", "买入手续费"),
    "exit_fee": ("exit_fee", "close_fee", "平仓手续费", "卖出手续费"),
    "notes": ("notes", "note", "review", "备注", "复盘", "总结"),
    "setup": ("setup", "pattern", "策略", "形态", "入场形态"),
    "market_regime": (
        "market_regime",
        "regime",
        "市场环境",
        "行情环境",
    ),
    "tags": ("tags", "tag", "标签"),
}

_KLINE_ALIASES = {
    "group_id": (
        "pattern_id",
        "shape_id",
        "trade_id",
        "group_id",
        "形态编号",
        "交易编号",
        "分组编号",
    ),
    "pattern_name": (
        "pattern_name",
        "shape_name",
        "name",
        "形态名称",
        "形态",
    ),
    "symbol": ("symbol", "ticker", "code", "股票", "股票代码", "标的"),
    "close_time": (
        "close_time",
        "close_time_ms",
        "timestamp",
        "time",
        "datetime",
        "date",
        "收盘时间",
        "时间",
    ),
    "open_time": ("open_time", "open_time_ms", "开盘时间"),
    "open": ("open", "开盘", "开盘价"),
    "high": ("high", "最高", "最高价"),
    "low": ("low", "最低", "最低价"),
    "close": ("close", "收盘", "收盘价"),
    "volume": ("volume", "vol", "成交量"),
    "interval": ("interval", "timeframe", "周期", "时间周期"),
}


def default_experience_path() -> Path:
    return default_config_path().with_name("external_trade_experiences.json")


def import_external_experiences(
    *,
    trade_path: Path | None = None,
    kline_path: Path | None = None,
    pattern_bars: int = 20,
) -> ExperienceImportResult:
    if trade_path is None and kline_path is None:
        raise ExperienceError("请至少选择一个交易记录或K线形态文件")
    if not 5 <= pattern_bars <= 240:
        raise ExperienceError("K线形态窗口必须在 5 到 240 根之间")

    trades = load_external_trades(trade_path) if trade_path is not None else []
    bars = load_external_klines(kline_path) if kline_path is not None else []
    used_groups: set[str] = set()
    if bars and trades:
        trades, used_groups = _attach_trade_patterns(trades, bars, pattern_bars)
    patterns = _standalone_pattern_records(
        bars,
        window=pattern_bars,
        include_ungrouped=not trades,
        excluded_groups=used_groups,
    )
    experiences = sorted(
        [*trades, *patterns],
        key=lambda item: (item.entry_time_ms, item.experience_id),
    )
    return ExperienceImportResult(
        experiences=experiences,
        trade_rows=len(trades),
        kline_rows=len(bars),
    )


def load_external_trades(path: Path) -> list[TradeExperience]:
    rows = _load_table(Path(path))
    experiences: list[TradeExperience] = []
    seen_ids: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        try:
            symbol = _required_text(row, _TRADE_ALIASES["symbol"], "symbol").upper()
            side = _normalize_side(_optional_text(row, _TRADE_ALIASES["side"]))
            entry_time_ms = _parse_timestamp_ms(
                _required_value(row, _TRADE_ALIASES["entry_time"], "entry_time")
            )
            exit_time_ms = _parse_timestamp_ms(
                _required_value(row, _TRADE_ALIASES["exit_time"], "exit_time")
            )
            if exit_time_ms < entry_time_ms:
                raise ValueError("平仓时间早于开仓时间")
            entry_price = _positive_decimal(
                _required_value(row, _TRADE_ALIASES["entry_price"], "entry_price"),
                "entry_price",
            )
            exit_price = _positive_decimal(
                _required_value(row, _TRADE_ALIASES["exit_price"], "exit_price"),
                "exit_price",
            )
            quantity = _positive_decimal(
                _required_value(row, _TRADE_ALIASES["quantity"], "quantity"),
                "quantity",
            )
            total_fee = _nonnegative_decimal(
                _optional_value(row, _TRADE_ALIASES["fee"]), "fee"
            )
            entry_fee_raw = _optional_value(row, _TRADE_ALIASES["entry_fee"])
            exit_fee_raw = _optional_value(row, _TRADE_ALIASES["exit_fee"])
            entry_fee = _nonnegative_decimal(entry_fee_raw, "entry_fee")
            exit_fee = _nonnegative_decimal(exit_fee_raw, "exit_fee")
            if entry_fee_raw in (None, "") and exit_fee_raw in (None, ""):
                exit_fee = total_fee
            elif total_fee > 0:
                raise ValueError("fee不能与entry_fee/exit_fee同时使用")

            sign = Decimal("1") if side == "LONG" else Decimal("-1")
            gross_pnl = (exit_price - entry_price) * quantity * sign
            net_pnl = gross_pnl - entry_fee - exit_fee
            entry_notional = entry_price * quantity
            return_percent = net_pnl / entry_notional * Decimal("100")
            outcome = (
                "WIN" if net_pnl > 0 else "LOSS" if net_pnl < 0 else "BREAKEVEN"
            )
            supplied_id = _optional_text(row, _TRADE_ALIASES["trade_id"])
            external_id = supplied_id or _stable_id(
                symbol,
                side,
                str(entry_time_ms),
                str(exit_time_ms),
                str(entry_price),
                str(exit_price),
                str(quantity),
            )
            if external_id in seen_ids:
                raise ValueError(f"交易编号重复：{external_id}")
            seen_ids.add(external_id)
            notes = tuple(
                value
                for value in (
                    _optional_text(row, _TRADE_ALIASES["notes"]),
                    _optional_text(row, _TRADE_ALIASES["setup"]),
                    _optional_text(row, _TRADE_ALIASES["tags"]),
                )
                if value
            )
            market_regime = _optional_text(
                row, _TRADE_ALIASES["market_regime"]
            )
            experiences.append(
                TradeExperience(
                    experience_id=f"trade_{_stable_id(external_id)}",
                    record_type="TRADE",
                    external_id=external_id,
                    symbol=symbol,
                    side=side,
                    outcome=outcome,
                    entry_time_ms=entry_time_ms,
                    exit_time_ms=exit_time_ms,
                    entry_time=_iso_time(entry_time_ms),
                    exit_time=_iso_time(exit_time_ms),
                    quantity=_decimal_text(quantity),
                    entry_price=_decimal_text(entry_price),
                    exit_price=_decimal_text(exit_price),
                    entry_fee=_decimal_text(entry_fee),
                    exit_fee=_decimal_text(exit_fee),
                    gross_pnl=_decimal_text(gross_pnl),
                    net_pnl=_decimal_text(net_pnl),
                    return_percent=_decimal_text(return_percent, places=6),
                    holding_seconds=(exit_time_ms - entry_time_ms) // 1000,
                    source="external_trade_file",
                    market_context=(
                        {"market_regime": market_regime} if market_regime else {}
                    ),
                    notes=notes,
                )
            )
        except (ExperienceError, InvalidOperation, TypeError, ValueError) as exc:
            raise ExperienceError(f"交易记录第 {row_number} 行格式错误：{exc}") from exc
    return experiences


def load_external_klines(path: Path) -> list[MarketBar]:
    rows = _load_table(Path(path))
    bars: list[MarketBar] = []
    for row_number, row in enumerate(rows, start=2):
        try:
            symbol = _optional_text(row, _KLINE_ALIASES["symbol"]).upper()
            group_id = _optional_text(row, _KLINE_ALIASES["group_id"])
            if not symbol and not group_id:
                raise ValueError("symbol和pattern_id不能同时为空")
            symbol = symbol or "GENERIC"
            interval = _optional_text(row, _KLINE_ALIASES["interval"])
            close_time = _optional_value(row, _KLINE_ALIASES["close_time"])
            open_time = _optional_value(row, _KLINE_ALIASES["open_time"])
            if close_time not in (None, ""):
                timestamp_ms = _parse_timestamp_ms(close_time)
            elif open_time not in (None, ""):
                if not interval:
                    raise ValueError("使用open_time时必须提供interval")
                timestamp_ms = _parse_timestamp_ms(open_time) + _parse_interval_ms(
                    interval
                )
            else:
                raise ValueError("缺少close_time/timestamp或open_time")

            values = {
                name: _positive_decimal(
                    _required_value(row, _KLINE_ALIASES[name], name), name
                )
                for name in ("open", "high", "low", "close")
            }
            volume = _nonnegative_decimal(
                _optional_value(row, _KLINE_ALIASES["volume"]), "volume"
            )
            if values["high"] < max(values["open"], values["close"]):
                raise ValueError("最高价低于开盘价或收盘价")
            if values["low"] > min(values["open"], values["close"]):
                raise ValueError("最低价高于开盘价或收盘价")
            bars.append(
                MarketBar(
                    symbol=symbol,
                    timestamp_ms=timestamp_ms,
                    open=values["open"],
                    high=values["high"],
                    low=values["low"],
                    close=values["close"],
                    volume=volume,
                    interval=interval,
                    group_id=group_id,
                    pattern_name=_optional_text(
                        row, _KLINE_ALIASES["pattern_name"]
                    ),
                )
            )
        except (ExperienceError, InvalidOperation, TypeError, ValueError) as exc:
            raise ExperienceError(f"K线记录第 {row_number} 行格式错误：{exc}") from exc
    bars.sort(key=lambda item: (item.group_id, item.symbol, item.timestamp_ms))
    return bars


def summarize_experiences(
    experiences: Iterable[TradeExperience],
) -> ExperienceSummary:
    items = list(experiences)
    trades = [item for item in items if item.record_type == "TRADE"]
    return ExperienceSummary(
        total=len(items),
        trades=len(trades),
        patterns=sum(item.record_type == "KLINE_PATTERN" for item in items),
        wins=sum(item.outcome == "WIN" for item in trades),
        losses=sum(item.outcome == "LOSS" for item in trades),
        breakeven=sum(item.outcome == "BREAKEVEN" for item in trades),
        with_kline=sum(
            bool(item.pre_entry_pattern.get("available")) for item in items
        ),
        net_pnl=sum(
            (Decimal(item.net_pnl) for item in trades if item.net_pnl), Decimal("0")
        ),
    )


def extract_bar_pattern(
    bars: Iterable[MarketBar],
    *,
    window: int,
    cutoff_ms: int | None = None,
) -> dict[str, Any]:
    eligible = [
        bar for bar in bars if cutoff_ms is None or bar.timestamp_ms < cutoff_ms
    ]
    selected = sorted(eligible, key=lambda item: item.timestamp_ms)[-window:]
    if not selected:
        return {
            "available": False,
            "bar_count": 0,
            "reason": "没有符合时间条件的已收盘K线",
        }
    base = selected[0].close
    average_volume = sum((bar.volume for bar in selected), Decimal("0")) / Decimal(
        len(selected)
    )

    def pct(value: Decimal) -> str:
        return _decimal_text((value / base - Decimal("1")) * Decimal("100"), 6)

    normalized = [
        {
            "timestamp_ms": bar.timestamp_ms,
            "open_pct": pct(bar.open),
            "high_pct": pct(bar.high),
            "low_pct": pct(bar.low),
            "close_pct": pct(bar.close),
            "body_pct": _decimal_text(
                (bar.close - bar.open) / base * Decimal("100"), 6
            ),
            "range_pct": _decimal_text(
                (bar.high - bar.low) / base * Decimal("100"), 6
            ),
            "volume_ratio": _decimal_text(
                bar.volume / average_volume
                if average_volume > 0
                else Decimal("0"),
                6,
            ),
        }
        for bar in selected
    ]
    close_change = (selected[-1].close / selected[0].close - Decimal("1")) * Decimal(
        "100"
    )
    total_range = (
        max(bar.high for bar in selected) - min(bar.low for bar in selected)
    ) / base * Decimal("100")
    last_volume_ratio = (
        selected[-1].volume / average_volume
        if average_volume > 0
        else Decimal("0")
    )
    trend = "UP" if close_change >= 1 else "DOWN" if close_change <= -1 else "FLAT"
    volatility = "HIGH" if total_range >= 3 else "MEDIUM" if total_range >= 1 else "LOW"
    volume_state = (
        "EXPANDING"
        if last_volume_ratio >= Decimal("1.5")
        else "SHRINKING"
        if last_volume_ratio <= Decimal("0.67")
        else "NORMAL"
    )
    names = [bar.pattern_name for bar in selected if bar.pattern_name]
    return {
        "available": True,
        "bar_count": len(selected),
        "pattern_name": names[-1] if names else "",
        "interval": selected[-1].interval,
        "start_time_ms": selected[0].timestamp_ms,
        "end_time_ms": selected[-1].timestamp_ms,
        "close_change_pct": _decimal_text(close_change, 6),
        "total_range_pct": _decimal_text(total_range, 6),
        "last_volume_ratio": _decimal_text(last_volume_ratio, 6),
        "shape_signature": f"{trend}_{volatility}_{volume_state}",
        "normalized_bars": normalized,
    }


def write_experience_document(
    path: Path, experiences: Iterable[TradeExperience]
) -> Path:
    target = Path(path)
    payload = _experience_document(list(experiences))
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(target.suffix + ".tmp")
    try:
        temporary.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        temporary.replace(target)
    except OSError as exc:
        raise ExperienceError(f"写入经验文件失败：{exc}") from exc
    return target


def merge_experience_document(
    path: Path, experiences: Iterable[TradeExperience]
) -> tuple[Path, int, int]:
    target = Path(path)
    existing: dict[str, dict[str, Any]] = {}
    if target.exists():
        try:
            payload = json.loads(target.read_text(encoding="utf-8"))
            if not isinstance(payload, dict) or payload.get("schema_version") != 2:
                raise ValueError("经验库版本不兼容")
            rows = payload.get("experiences", [])
            if not isinstance(rows, list):
                raise ValueError("experiences不是列表")
            for row in rows:
                if not isinstance(row, dict):
                    raise ValueError("经验记录必须是对象")
                if row.get("source") not in {
                    "external_trade_file",
                    "external_kline_file",
                }:
                    raise ValueError("经验库包含非外部导入记录")
                experience_id = row.get("experience_id")
                if not isinstance(experience_id, str) or not experience_id:
                    raise ValueError("经验记录缺少experience_id")
                existing[experience_id] = row
        except (OSError, UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            raise ExperienceError(f"现有经验库格式错误：{exc}") from exc
    previous = len(existing)
    for item in experiences:
        if item.source not in {"external_trade_file", "external_kline_file"}:
            raise ExperienceError("只允许保存从外部文件导入的经验记录")
        existing[item.experience_id] = asdict(item)
    merged = sorted(
        existing.values(),
        key=lambda row: (
            int(row.get("entry_time_ms", 0)),
            str(row.get("experience_id", "")),
        ),
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(target.suffix + ".tmp")
    payload = {
        "schema_version": 2,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(merged),
        "experiences": merged,
    }
    try:
        temporary.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        temporary.replace(target)
    except OSError as exc:
        raise ExperienceError(f"写入本地经验库失败：{exc}") from exc
    return target, len(merged) - previous, len(merged)


class OpenAIVectorStoreUploader:
    def __init__(
        self,
        *,
        request_json: Callable[[str, dict[str, Any], str, int], dict[str, Any]]
        | None = None,
        upload_file: Callable[[Path, str, int], dict[str, Any]] | None = None,
    ) -> None:
        self._request_json = request_json or _post_json
        self._upload_file = upload_file or _upload_openai_file

    def upload(
        self,
        path: Path,
        *,
        api_key: str,
        vector_store_id: str = "",
        timeout_seconds: int = 30,
    ) -> UploadResult:
        if not api_key.strip():
            raise ExperienceError("缺少 OpenAI API Key")
        if not 5 <= timeout_seconds <= 120:
            raise ExperienceError("上传超时必须在 5 到 120 秒之间")
        store_id = vector_store_id.strip()
        if store_id and not _valid_resource_id(store_id, "vs_"):
            raise ExperienceError("Vector Store ID 格式不正确")
        if not store_id:
            store = self._request_json(
                OPENAI_VECTOR_STORES_URL,
                {"name": "AutoQuant 外部交易经验库"},
                api_key,
                timeout_seconds,
            )
            store_id = str(store.get("id", ""))
            if not _valid_resource_id(store_id, "vs_"):
                raise ExperienceError("OpenAI 未返回有效的 Vector Store ID")

        uploaded = self._upload_file(Path(path), api_key, timeout_seconds)
        file_id = str(uploaded.get("id", ""))
        if not _valid_resource_id(file_id, "file-"):
            raise ExperienceError("OpenAI 未返回有效的文件 ID")
        attached = self._request_json(
            f"{OPENAI_VECTOR_STORES_URL}/{store_id}/files",
            {"file_id": file_id},
            api_key,
            timeout_seconds,
        )
        return UploadResult(
            vector_store_id=store_id,
            file_id=file_id,
            attachment_id=str(attached.get("id", "")),
            status=str(attached.get("status", "in_progress")),
        )


def _attach_trade_patterns(
    trades: list[TradeExperience],
    bars: list[MarketBar],
    window: int,
) -> tuple[list[TradeExperience], set[str]]:
    grouped: dict[str, list[MarketBar]] = {}
    ungrouped: dict[str, list[MarketBar]] = {}
    for bar in bars:
        if bar.group_id:
            grouped.setdefault(bar.group_id, []).append(bar)
        else:
            ungrouped.setdefault(bar.symbol, []).append(bar)
    result: list[TradeExperience] = []
    used_groups: set[str] = set()
    for trade in trades:
        if trade.external_id in grouped:
            candidates = grouped[trade.external_id]
            symbols = {bar.symbol for bar in candidates if bar.symbol != "GENERIC"}
            if symbols and symbols != {trade.symbol}:
                raise ExperienceError(
                    f"K线分组 {trade.external_id} 的symbol与交易记录不一致"
                )
            used_groups.add(trade.external_id)
        else:
            candidates = ungrouped.get(trade.symbol, [])
        intervals = {bar.interval for bar in candidates if bar.interval}
        if len(intervals) > 1:
            raise ExperienceError(
                f"交易 {trade.external_id} 关联了多个K线周期，请使用编号分组"
            )
        result.append(
            replace(
                trade,
                pre_entry_pattern=extract_bar_pattern(
                    candidates, window=window, cutoff_ms=trade.entry_time_ms
                ),
            )
        )
    return result, used_groups


def _standalone_pattern_records(
    bars: list[MarketBar],
    *,
    window: int,
    include_ungrouped: bool,
    excluded_groups: set[str],
) -> list[TradeExperience]:
    groups: dict[str, list[MarketBar]] = {}
    for bar in bars:
        if bar.group_id:
            if bar.group_id not in excluded_groups:
                groups.setdefault(bar.group_id, []).append(bar)
        elif include_ungrouped:
            key = f"{bar.symbol}::{bar.interval or 'unknown'}"
            groups.setdefault(key, []).append(bar)

    result: list[TradeExperience] = []
    for group_id, group_bars in sorted(groups.items()):
        selected = sorted(group_bars, key=lambda item: item.timestamp_ms)[-window:]
        if not selected:
            continue
        pattern = extract_bar_pattern(selected, window=window)
        external_id = group_bars[0].group_id or group_id
        pattern_name = str(pattern.get("pattern_name", ""))
        explicit_symbols = {
            bar.symbol for bar in group_bars if bar.symbol != "GENERIC"
        }
        if len(explicit_symbols) > 1:
            raise ExperienceError(f"K线分组 {external_id} 包含多个symbol")
        intervals = {bar.interval for bar in group_bars if bar.interval}
        if len(intervals) > 1:
            raise ExperienceError(f"K线分组 {external_id} 包含多个interval")
        symbol = next(iter(explicit_symbols), selected[-1].symbol)
        result.append(
            TradeExperience(
                experience_id=f"pattern_{_stable_id(external_id)}",
                record_type="KLINE_PATTERN",
                external_id=external_id,
                symbol=symbol,
                side="NONE",
                outcome="UNLABELED",
                entry_time_ms=selected[0].timestamp_ms,
                exit_time_ms=selected[-1].timestamp_ms,
                entry_time=_iso_time(selected[0].timestamp_ms),
                exit_time=_iso_time(selected[-1].timestamp_ms),
                quantity="",
                entry_price=_decimal_text(selected[0].close),
                exit_price=_decimal_text(selected[-1].close),
                entry_fee="",
                exit_fee="",
                gross_pnl="",
                net_pnl="",
                return_percent=str(pattern.get("close_change_pct", "")),
                holding_seconds=max(
                    0, (selected[-1].timestamp_ms - selected[0].timestamp_ms) // 1000
                ),
                source="external_kline_file",
                pre_entry_pattern=pattern,
                notes=(pattern_name,) if pattern_name else (),
            )
        )
    return result


def _load_table(path: Path) -> list[dict[str, Any]]:
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise ExperienceError(f"无法读取导入文件：{exc}") from exc
    if size > MAX_IMPORT_FILE_BYTES:
        raise ExperienceError("单个导入文件不能超过 50 MB")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv_rows(path)
    if suffix == ".xlsx":
        return _load_xlsx_rows(path)
    raise ExperienceError("只支持 .xlsx 或 .csv 文件")


def _load_csv_rows(path: Path) -> list[dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise ExperienceError("CSV缺少表头")
            headers = [_normalize_header(value) for value in reader.fieldnames]
            if any(not header for header in headers):
                raise ExperienceError("CSV表头不能为空")
            if len(set(headers)) != len(headers):
                raise ExperienceError("CSV表头不能重复")
            reader.fieldnames = headers
            rows = []
            for index, row in enumerate(reader, start=1):
                if index > MAX_IMPORT_ROWS:
                    raise ExperienceError("导入文件不能超过 500000 行")
                if None in row:
                    raise ExperienceError(f"CSV第 {index + 1} 行列数超过表头")
                if not any(value not in (None, "") for value in row.values()):
                    continue
                rows.append(row)
            return rows
    except UnicodeDecodeError as exc:
        raise ExperienceError("CSV必须使用 UTF-8 编码") from exc
    except OSError as exc:
        raise ExperienceError(f"读取CSV失败：{exc}") from exc


def _load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExperienceError("读取Excel需要安装 openpyxl") from exc
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=True, keep_links=False
        )
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        if not header_values:
            raise ExperienceError("Excel首个工作表缺少表头")
        headers = [_normalize_header(value) for value in header_values]
        if any(not header for header in headers):
            raise ExperienceError("Excel表头不能为空")
        if len(set(headers)) != len(headers):
            raise ExperienceError("Excel表头不能重复")
        rows = []
        for index, values in enumerate(iterator, start=1):
            if index > MAX_IMPORT_ROWS:
                raise ExperienceError("导入文件不能超过 500000 行")
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(
                {
                    header: values[position] if position < len(values) else None
                    for position, header in enumerate(headers)
                }
            )
        return rows
    except ExperienceError:
        raise
    except (OSError, ValueError) as exc:
        raise ExperienceError(f"读取Excel失败：{exc}") from exc
    finally:
        if "workbook" in locals():
            workbook.close()


def _experience_document(items: list[TradeExperience]) -> dict[str, Any]:
    return {
        "schema_version": 2,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(items),
        "experiences": [asdict(item) for item in items],
    }


def _upload_openai_file(
    path: Path, api_key: str, timeout_seconds: int
) -> dict[str, Any]:
    try:
        content = path.read_bytes()
    except OSError as exc:
        raise ExperienceError(f"无法读取待上传文件：{exc}") from exc
    boundary = f"----AutoQuant{uuid.uuid4().hex}"
    filename = path.name.encode("ascii", "ignore").decode("ascii") or "experiences.json"
    chunks = [
        f"--{boundary}\r\n".encode(),
        b'Content-Disposition: form-data; name="purpose"\r\n\r\n',
        b"assistants\r\n",
        f"--{boundary}\r\n".encode(),
        (
            'Content-Disposition: form-data; name="file"; '
            f'filename="{filename}"\r\n'
        ).encode(),
        b"Content-Type: application/json\r\n\r\n",
        content,
        b"\r\n",
        f"--{boundary}--\r\n".encode(),
    ]
    request = Request(
        OPENAI_FILES_URL,
        data=b"".join(chunks),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Accept": "application/json",
            "User-Agent": "AutoQuant/0.5.0",
        },
        method="POST",
    )
    return _read_json_request(request, timeout_seconds)


def _post_json(
    url: str,
    payload: dict[str, Any],
    api_key: str,
    timeout_seconds: int,
) -> dict[str, Any]:
    request = Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "AutoQuant/0.5.0",
        },
        method="POST",
    )
    return _read_json_request(request, timeout_seconds)


def _read_json_request(request: Request, timeout_seconds: int) -> dict[str, Any]:
    try:
        with urlopen(request, timeout=timeout_seconds) as response:
            content = response.read(MAX_HTTP_RESPONSE_BYTES + 1)
    except HTTPError as exc:
        message = f"OpenAI HTTP {exc.code}"
        try:
            body = exc.read(16_384)
            payload = json.loads(body.decode("utf-8"))
            error = payload.get("error") if isinstance(payload, dict) else None
            detail = error.get("message") if isinstance(error, dict) else None
            if isinstance(detail, str) and detail.strip():
                message += "：" + " ".join(detail.split())[:300]
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            pass
        raise ExperienceError(message) from exc
    except (URLError, TimeoutError, OSError) as exc:
        raise ExperienceError("OpenAI 上传网络请求失败或超时") from exc
    if len(content) > MAX_HTTP_RESPONSE_BYTES:
        raise ExperienceError("OpenAI 上传响应超过大小限制")
    try:
        result = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ExperienceError("OpenAI 返回了无法解析的上传响应") from exc
    if not isinstance(result, dict):
        raise ExperienceError("OpenAI 上传响应格式错误")
    return result


def _required_value(
    row: dict[str, Any], aliases: tuple[str, ...], field_name: str
) -> Any:
    value = _optional_value(row, aliases)
    if value in (None, ""):
        raise ValueError(f"缺少{field_name}")
    return value


def _required_text(
    row: dict[str, Any], aliases: tuple[str, ...], field_name: str
) -> str:
    value = _optional_text(row, aliases)
    if not value:
        raise ValueError(f"缺少{field_name}")
    return value


def _optional_value(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        value = row.get(_normalize_header(alias))
        if value not in (None, ""):
            return value
    return None


def _optional_text(row: dict[str, Any], aliases: tuple[str, ...]) -> str:
    value = _optional_value(row, aliases)
    return " ".join(str(value).split())[:500] if value not in (None, "") else ""


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_side(value: str) -> str:
    text = value.strip().upper()
    if text in {"", "LONG", "BUY", "多", "做多", "买入"}:
        return "LONG"
    if text in {"SHORT", "SELL", "空", "做空", "卖出"}:
        return "SHORT"
    raise ValueError("side必须是LONG/SHORT、BUY/SELL或多/空")


def _positive_decimal(value: Any, field_name: str) -> Decimal:
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name}必须是数字") from exc
    if not number.is_finite() or number <= 0:
        raise ValueError(f"{field_name}必须是正数")
    return number


def _nonnegative_decimal(value: Any, field_name: str) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name}必须是数字") from exc
    if not number.is_finite() or number < 0:
        raise ValueError(f"{field_name}不能是负数")
    return number


def _parse_timestamp_ms(value: Any) -> int:
    if isinstance(value, datetime):
        parsed = value
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return int(parsed.timestamp() * 1000)
    if isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
        return int(parsed.timestamp() * 1000)
    text = str(value).strip()
    if not text:
        raise ValueError("时间不能为空")
    try:
        numeric = Decimal(text)
        if not numeric.is_finite() or numeric <= 0:
            raise ValueError("时间戳必须为正数")
        result = int(numeric)
        if result < 100_000_000_000:
            result *= 1000
        return result
    except InvalidOperation:
        pass
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError("时间必须是毫秒、秒或ISO时间") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return int(parsed.timestamp() * 1000)


def _parse_interval_ms(value: str) -> int:
    text = value.strip().lower()
    if len(text) < 2 or text[-1] not in {"s", "m", "h", "d"}:
        raise ValueError("interval必须类似1m、5m、1h或1d")
    try:
        amount = int(text[:-1])
    except ValueError as exc:
        raise ValueError("interval必须类似1m、5m、1h或1d") from exc
    if amount <= 0:
        raise ValueError("interval必须为正数")
    multiplier = {"s": 1_000, "m": 60_000, "h": 3_600_000, "d": 86_400_000}
    return amount * multiplier[text[-1]]


def _valid_resource_id(value: str, prefix: str) -> bool:
    if not value.startswith(prefix) or len(value) > 200:
        return False
    return all(character.isalnum() or character in "_-" for character in value)


def _stable_id(*parts: str) -> str:
    content = "\x1f".join(parts).encode("utf-8")
    return hashlib.sha256(content).hexdigest()[:24]


def _decimal_text(value: Decimal, places: int | None = None) -> str:
    normalized = value
    if places is not None:
        quantum = Decimal("1").scaleb(-places)
        normalized = value.quantize(quantum)
    return format(normalized, "f")


def _iso_time(timestamp_ms: int) -> str:
    return datetime.fromtimestamp(timestamp_ms / 1000, timezone.utc).isoformat()
